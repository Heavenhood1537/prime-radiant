"""Regenerate both master XLSX files from milestones.json (run manually).

Run after ANY dataset change, together with rebuilding seldon.db:
    python seldon_gaps.py        # optional, gap report
    python regen_xlsx.py         # refresh both Excel masters
    python build_seldon_db.py    # re-embed (stop seldon_server.py first)

Outputs:
  - MILESTONES COMPLETE LATEST - RESTORED.xlsx  (project copy, this folder)
  - C:\\Users\\milan\\Documents\\1_Desktop\\RADIANT\\MILESTONES WORK\\
    MILESTONES DATASETS\\MILESTONES COMPLETE LATEST.xlsx  (master copy)

Both are full chronological rebuilds - 9 columns matching the original
master schema, with a Years/Ago formula so the sheet stays self-updating.
"""
import json
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_ROOT = os.path.dirname(os.path.abspath(__file__))
JSON_PATH = os.path.join(_ROOT, "milestones.json")
OUT_PROJECT = os.path.join(_ROOT, "MILESTONES COMPLETE LATEST - RESTORED.xlsx")
OUT_MASTER = (r"C:\Users\milan\Documents\1_Desktop\RADIANT"
              r"\MILESTONES WORK\MILESTONES DATASETS"
              r"\MILESTONES COMPLETE LATEST.xlsx")
CURRENT_YEAR = 2026

HEADERS = ["Category", "Originating Idea", "Latitude", "Longitude",
           "Time/Period", "Years/Ago", "Originator", "Description", "DetailsURL"]
WIDTHS = [22, 55, 11, 11, 13, 12, 38, 90, 45]


def fmt_year(y):
    if y is None:
        return ""
    return f"{-y} BCE" if y < 0 else f"{y} CE"


def build(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Milestones"
    fill = PatternFill("solid", fgColor="1F3864")
    for c, (h, w) in enumerate(zip(HEADERS, WIDTHS), start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 22
    for r, m in enumerate(rows, start=2):
        loc = m.get("location") or {}
        ws.cell(row=r, column=1, value=m.get("category"))
        ws.cell(row=r, column=2, value=m.get("title"))
        ws.cell(row=r, column=3, value=loc.get("lat"))
        ws.cell(row=r, column=4, value=loc.get("lon"))
        ws.cell(row=r, column=5, value=m.get("year"))
        ws.cell(row=r, column=6, value=f"={CURRENT_YEAR}-E{r}")
        ws.cell(row=r, column=7, value="; ".join(m.get("originators") or []))
        ws.cell(row=r, column=8, value=m.get("description"))
        ws.cell(row=r, column=9, value=m.get("url"))
    last = len(rows) + 1
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:I{last}"
    return wb


rows = sorted(json.load(open(JSON_PATH, encoding="utf-8"))["milestones"],
              key=lambda m: (m["year"] is None, m["year"] or 0))
for path in (OUT_PROJECT, OUT_MASTER):
    build(rows).save(path)
    print(f"saved: {path}  ({len(rows)} rows)")
